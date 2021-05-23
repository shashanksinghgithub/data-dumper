
import xlwings as xw
import numpy as np
import pandas as pd
from collections import OrderedDict
#from time import time
from dateutil import parser
#import os
import openpyxl
#import sys
import logging
#import Email_func
#from forex_python.converter import CurrencyRates
import shutil
import os 
import datetime

master_dir = 'X:\\Data_dumpers\\ETL\\Master'
download_dir = 'X:\\Data_dumpers\\ETL\\Download\\'
output_dir = 'X:\\Data_dumpers\\ETL\\Output\\'
log_path = "X:\\Data_dumpers\\ETL\\"
email_dir = "X:\\Data_dumpers\\ETL\\Output\\"
market_snapshot = "X:\\Market_Snapshot\\"

logging.basicConfig(filename=log_path+"test.log",
                        level=logging.DEBUG,
                        format="%(asctime)s:%(levelname)s:%(message)s")


# calculate SSF roll over

def ssf_calculator(a, b):
    try:
        return round(b / (a + b), 3)
    except:
        return 0.0


def bhavcopy_processor(file_name):
    '''Function to read and process bhavcopy
    Convert it into a pivot_table
    calculate product-far/ product-near and return that dataframe '''

    # read bhav copy as workbook
    logging.info('Reading bhavcopy')
    app = xw.App(visible=False)
    
    wb = app.books.open(download_dir + file_name)

    # read sheet
    sht = wb.sheets[file_name[:-4]]

    # convert sheet to pandas dataframe
    all_stocks = sht.range('A1').options(pd.DataFrame, expand='table').value

    # select only FUTSTK and FUTIDX
    
    new_index = ['FUTSTK', 'FUTIDX']
    try:
        logging.info('Selecting FUTSTK and FUTIDX')
        all_stocks = all_stocks.loc[new_index]
        
    except:
        logging.info('Not able to select FUTSTK and FUTIDX')

    
    # Categorization based on master document
    # read master doc
    logging.info('Reading master files...')
    
    mast_book = xw.Book(master_dir + 'MasterData.xlsx')

    # read sheet
    mast_sheet = mast_book.sheets('Sheet1')

    # convert sheet to pandas dataframe
    master = mast_sheet.range('A1').options(pd.DataFrame, expand='table').value
    # master.rename(columns={'SymbolName': 'SYMBOL'}, inplace=True)

    # filter master for symbol, bloomcode and sector
    master.set_index(keys='SYMBOL', inplace=True)

    master = master[master['IsActiveFNO']==True]  # filter on active fno stocks 
    master.drop(columns=['IsActiveFNO'], axis=1, inplace=True)

    # join dataframes master and all_stocks based on SYMBOL
    all_stocks = all_stocks.join(master, on='SYMBOL')
    #ignore fut weekly
    monthly_exp = all_stocks[all_stocks['Type']=='SSF']['EXPIRY_DT'].unique()
    all_stocks = all_stocks[all_stocks['EXPIRY_DT'].isin(monthly_exp)]
    
    # create Pivot table
    all_stocks = all_stocks.pivot_table(values=['OPEN_INT', 'SETTLE_PR'], index=['SYMBOL', 'Sector', 'IsNifty', 'Type'],
                                        columns='EXPIRY_DT', aggfunc={'SETTLE_PR': np.sum,
                                                                      'OPEN_INT': np.sum})

    # check for m2 and m3 months null values; if stocks go out of FNO
    all_stocks['OPEN_INT'].iloc[:,1].fillna(0, inplace=True)
    all_stocks['OPEN_INT'].iloc[:,2].fillna(0, inplace=True)
    
    all_stocks['SETTLE_PR'].iloc[:,1].fillna(0, inplace=True)
    all_stocks['SETTLE_PR'].iloc[:,2].fillna(0, inplace=True)    
    
    # calculate product-near and product-far
    data = pd.DataFrame([map(round, all_stocks['OPEN_INT'].iloc[:, 0] * all_stocks['SETTLE_PR'].iloc[:, 0])
                            , map(round, all_stocks['OPEN_INT'].iloc[:, 1] * all_stocks['SETTLE_PR'].iloc[:, 1] +
                                  all_stocks['OPEN_INT'].iloc[:, 2] * all_stocks['SETTLE_PR'].iloc[:, 2])]).T

    data.columns = ['Product-Near', 'Product-Far']
    data.index = all_stocks.index
    
    
    # apply roll over to all stocks and append ssf_rollover
    data['SSF_rollover'] = data.apply(lambda row: ssf_calculator(row['Product-Near'], row['Product-Far']), axis=1)
    pd.set_option('display.precision', 10)
    print 'Bhavcopy processed successfully...'
    logging.info('Bhavcopy processed successfully...')
    wb.close()
    
    return data, master


def sector_rollovers(data):
    '''Function to calculate sector vice rollovers and dump to base Report_stats.xlsx'''

    # Group all sectors and run sum() aggregate
    Sector = data.groupby('Sector').sum()
    Sector['Roll'] = Sector.apply(lambda row: ssf_calculator(row['Product-Near'], row['Product-Far']), axis=1)
    Sector.drop(columns='SSF_rollover', inplace=True, axis=1)

    # append nifty stocks to Sector table
    nifty_product_near = data.groupby('IsNifty').get_group(True).sum()['Product-Near']
    nifty_product_far = data.groupby('IsNifty').get_group(True).sum()['Product-Far']
    nifty_roll = ssf_calculator(nifty_product_near, nifty_product_far)

    Nifty_Stocks_df = pd.DataFrame(
        {'Product-Near': [nifty_product_near], 'Product-Far': [nifty_product_far], 'Roll': [nifty_roll]},
        index=['Nifty Stocks'])

    # append non nifty stocks to sector table
    non_nifty_product_near = data.groupby('IsNifty').get_group(False).groupby('Type').sum()['Product-Near']['SSF']
    non_nifty_product_far = data.groupby('IsNifty').get_group(False).groupby('Type').sum()['Product-Far']['SSF']
    non_nifty_roll = ssf_calculator(non_nifty_product_near, non_nifty_product_far)

    Non_Nifty_Stocks_df = pd.DataFrame(
        {'Product-Near': [non_nifty_product_near], 'Product-Far': [non_nifty_product_far], 'Roll': [non_nifty_roll]},
        index=['Non-Nifty Stocks'])

    # append SSF stock to sector table
    ssf_df = pd.DataFrame({'Product-Near': data.groupby('Type').sum().loc['SSF']['Product-Near'],
                           'Product-Far': data.groupby('Type').sum().loc['SSF']['Product-Far'],
                           'Roll': ssf_calculator(data.groupby('Type').sum().loc['SSF']['Product-Near'],
                                                  data.groupby('Type').sum().loc['SSF']['Product-Far'])},
                          index=['SSF'])

    # append NIFTY to sector table
    NIFTY_df = pd.DataFrame({'Product-Near': data.loc['NIFTY'].loc[:, 'Product-Near'][0],
                             'Product-Far': data.loc['NIFTY'].loc[:, 'Product-Far'][0],
                             'Roll': ssf_calculator(data.loc['NIFTY'].loc[:, 'Product-Near'][0],
                                                    data.loc['NIFTY'].loc[:, 'Product-Far'][0])},
                            index=['Nifty'])

    # append BANKNIFTY to sector table
    BANKNIFTY_df = pd.DataFrame({'Product-Near': data.loc['BANKNIFTY'].loc[:, 'Product-Near'][0],
                                 'Product-Far': data.loc['BANKNIFTY'].loc[:, 'Product-Far'][0],
                                 'Roll': ssf_calculator(data.loc['BANKNIFTY'].loc[:, 'Product-Near'][0],
                                                        data.loc['BANKNIFTY'].loc[:, 'Product-Far'][0])},
                                index=['NiftyBank'])

    Overall = pd.DataFrame({'Product-Near': data['Product-Near'].sum(),
                            'Product-Far': data['Product-Far'].sum(),
                            'Roll': round(
                                data['Product-Far'].sum() / (data['Product-Near'].sum() + data['Product-Far'].sum()),
                                2)},
                           index=['Overall'])

    frames = [Sector, Nifty_Stocks_df, Non_Nifty_Stocks_df, ssf_df, NIFTY_df, BANKNIFTY_df, Overall]
    result = pd.concat(frames, sort=False)
    result.sort_index(inplace=True)
    print result.columns
    #result.drop(index=['None', 'Index'], axis=0, inplace=True)
    result.drop(index=['Index'], axis=0, inplace=True)

    # write dataframe to excel
    writer = pd.ExcelWriter(output_dir + 'Report_stats.xlsx', engine='xlsxwriter')
    # COPY power_1turb (id, ts, value) FROM 'data' WITH DATETIMEFORMAT = '%Y-%m-%dT%H:%M:%SZ';
    result.to_excel(writer, sheet_name='SSF & Indices Report')
    logging.info('Successfully dumped sector vice rollovers to Report_stats file...')
    return result


def rollover_data_dumper(data, master, result, expiry, month, year, file_name, dollar_value):
    '''Function to dump rollover data for expiry'''
    logging.info('Dumping data for {} {} {}.....'.format(expiry, month, year))

    # ignore index i.e consider only SSF stocks

    ssf_rollover = data.groupby('Type').get_group('SSF').reset_index(level=[1, 2, 3], drop=True)[['SSF_rollover']]

    # fetch bloomcode from master and join dataframes to dumper

    stock_bloom = master.groupby('Type').get_group('SSF')[['BloomCode']]
    dumper = stock_bloom.merge(ssf_rollover, on='SYMBOL')

    # dump expiry, month and year

    dumper['Expiry'], dumper['Month'], dumper['Year'] = expiry, month, year
    dumper.reset_index(inplace=True)

    # reaarange according to schema
    
    dumper = dumper[['Expiry', 'SYMBOL', 'BloomCode', 'Month', 'Year', 'SSF_rollover']]
    writer = pd.ExcelWriter(output_dir + 'Report.xlsx', engine='xlsxwriter')
    
    dumper.to_excel(writer, sheet_name='SSF_rollovers', index=False)
    logging.info('SSF rollovers successfully dumped...')

    # dump indices rollover

    indices_rollover = pd.DataFrame(result['Roll'])
    indices_rollover['Expiry'], indices_rollover['Month'], indices_rollover['Year'] = expiry, month, year
    indices_rollover.index.name = 'Indices'
    indices_rollover.reset_index(inplace=True)
    indices_rollover = indices_rollover[['Expiry', 'Month', 'Year', 'Indices', 'Roll']]
    indices_rollover['Indices'] =  indices_rollover['Indices'].replace({'Technology':'tech','PSUBANKS':'PSU Banks'})
    
    indices_rollover.to_excel(writer, sheet_name='Indices_rollovers', index=False)
    logging.info('Indices rollover successfully dumped...')
    # Unrolled data

    unrolled = master.groupby('Type').get_group('SSF')[['BloomCode']].merge(
        data.groupby('Type').get_group('SSF').reset_index(
            level=[1, 2, 3], drop=True)[['Product-Near']], on='SYMBOL')

    
    try:
        #logging.info('Fetching dollar rate for {0}...'.format(file_name))
        #dollar_value = dollar.loc[file_name]['Rate'][0]
        #c = CurrencyRates()
        #dollar_value= c.get_rate('USD', 'INR')
        #dollar_value= 68.98
        unrolled['Million_dollar'] = unrolled['Product-Near'].apply(lambda x: round(x / (dollar_value * 1000000), 1))
        logging.info('Dollar rate for {0} is {1}-INR'.format(file_name, dollar_value))
        
    except:
        logging.exception('Dollar rate for {0} not found in master file...'.format(file_name))

    # dump million dollar rates of stocks

    unrolled_dumper = master.groupby('Type').get_group('SSF')[['BloomCode']]
    unrolled_dumper['Expiry'], unrolled_dumper['Month'], unrolled_dumper['Year'] = expiry, month, year
    unrolled_dumper['Million_dollar'] = unrolled['Million_dollar']
    unrolled_dumper.reset_index(inplace=True)
    unrolled_dumper = unrolled_dumper[['Expiry', 'SYMBOL', 'BloomCode', 'Month', 'Year', 'Million_dollar']]

    unrolled_dumper.to_excel(writer, sheet_name='Unrolled_rollovers', index=False)
    logging.info('Unrolled SSF rollovers successfully dumped...')
    writer.save()
    
    return dumper, indices_rollover, unrolled_dumper


# Calculate avergae of last three months
def average_calculator(expiry, value, stock_names, data_df, date):
    '''Function to calculate average of last three months'''
   
    #months.keys()[months.values().index(months[month]-1)]
    average = []
    dates = sorted(set(data_df['Date']), reverse=True)
    if date in dates:
        dates.remove(date)
    
    
    for stock in stock_names:
        
        try:
            average.append(round(data_df[(data_df['Expiry']==expiry) &
                                         ((data_df['Date']==dates[0]) |
                                                 (data_df['Date']==dates[1]) |
                                                 (data_df['Date']==dates[2])) ].loc[stock][value].mean(), 2) )  
        except:
            average.append(0)
    
      

    return average


# Sort elements based on criteria top or low performing stocks
def sort_stocks(rank_order, sorting_type, data, criteria):
    sorted_stocks = {}

    for group_name, group_elements in data.groupby('TYPE'):
        sorted_stocks[group_name] = group_elements.sort_values(ascending=sorting_type, by=criteria)[: rank_order]

    return sorted_stocks


def greaterthanAverage(expiry, month, year):
    '''Function for greater than average sheet'''

    # read from dumper object for SSF rolldata
    rolldata = pd.read_excel(master_dir + 'RollData.xlsx')
    rolldata = rolldata.set_index('SYMBOL')

    # average is appended in sorted manner
    stock_names = list(sorted(set(rolldata.index)))
    
    rolldata['Date'] = rolldata.apply(lambda row: pd.to_datetime(row['Month']+' '+str(row['Year']) ), axis=1)
    date = pd.to_datetime(month+" "+year)
    
    logging.info('Calculating average ssf rollovers for last three months...')
    average_ssf = pd.DataFrame(average_calculator(expiry, 'SSF_rollover', stock_names, rolldata, date), index=stock_names,
                               columns=['Average'])
    
    average_ssf.index.name = 'SYMBOL'
    average_ssf = average_ssf[average_ssf['Average']!=0]
    
    # consolidate data to already exisiting file
    from openpyxl import load_workbook
    book = load_workbook(output_dir+'Report.xlsx')
    
    writer = pd.ExcelWriter(output_dir+'Report.xlsx', engine='openpyxl') 
    writer.book = book
    
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    average_ssf.to_excel(writer, "SSF_avearge")
    writer.save()
        
    
    #average_ssf.to_excel(output_dir+'SSF_average.xlsx')
    # Greater than average
    # read from dumped data
    grtthanaverage = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='SSF_rollovers')

    grtthanaverage = grtthanaverage.merge(average_ssf, on='SYMBOL')

    # difference of today's rollover and average of last 3 months
    grtthanaverage['Difference'] = grtthanaverage['SSF_rollover'] - grtthanaverage['Average']

    # read master for fetching type of stock nifty, non-nifty and physical
    type_of_stock = pd.read_excel(master_dir + 'MasterData.xlsx')
    type_of_stock = type_of_stock[['SYMBOL','IsNifty','IsActiveFNO']]
    type_of_stock = type_of_stock[type_of_stock['IsActiveFNO']==True][['SYMBOL','IsNifty']]
    type_of_stock['TYPE'] = np.where(type_of_stock['IsNifty']==True, 'Nifty', 'Non Nifty')
    type_of_stock = type_of_stock[['SYMBOL','TYPE']]
    
    
    grtthanaverage = grtthanaverage.merge(type_of_stock, on='SYMBOL')
    # fetch symbols that are active FNO
    #active_fno = pd.read_excel(master_dir + "MasterData.xlsx")
    #active_fno = active_fno[active_fno['IsActiveFNO']==]
    grtthanaverage.to_csv("debug.csv")
   
    # pass False for top performing as sorting parameter
    logging.info('Fetching top 5 stocks...')
    grt_avg_top_stocks = sort_stocks(5, False, grtthanaverage, 'Difference')   
    
    
    for key, values in grt_avg_top_stocks.items():      
        values = values[values['Difference']>0]
        if len(values)!=0:
            grt_avg_top_stocks[key] = values.sort_values(ascending=False, by=['SSF_rollover']).head(5)
        else:
            print "No stocks with greater than 3 month average"
            grt_avg_top_stocks[key] = 'None' 
            
        
    # pass True for low performing as sorting parameter
    logging.info('Fetching bottom 5 stocks...')
    grt_avg_low_stocks = sort_stocks(5, True, grtthanaverage, 'Difference')
    
    for key, values in grt_avg_low_stocks.items():
        values = values[values['Difference']<0]
        if len(values)!=0:
            grt_avg_low_stocks[key] = values.sort_values(ascending= True, by=['SSF_rollover']).head(5)
        else:
            print "No stocks with lower than 3 month average"
            grt_avg_low_stocks[key] = 'None'
        
 
        
    return grt_avg_top_stocks, grt_avg_low_stocks


def del_qty_average():
    '''Function to calculate last 3months del quantity average'''

    # read del qty dump
    del_qty = pd.read_excel(master_dir + 'Del Qty.xlsx')
    # parse del qty dates to date format
    del_qty['Date'] = del_qty.apply(lambda row: parser.parse(row['Date']), axis=1)
    del_qty.set_index('SYMBOL', inplace=True)
    stock_names = pd.read_excel(master_dir+"MasterData.xlsx")
    stock_names = list(stock_names[ (stock_names['IsActiveFNO']==True)  & (stock_names['Type']=='SSF') ]['SYMBOL'].values)

    
    average = {}
    for name in stock_names:
        try:
            average[name] = round(del_qty.loc[name].sort_values(ascending=False, by='Date').iloc[:3, -1].mean(),2)
        except:
            print "New code added {}".format(name)

    average_del = pd.DataFrame(average.items(), columns=['SYMBOL', 'Del Avg'])
    average_del.dropna(inplace=True, axis=0)
    average_del.set_index('SYMBOL', inplace=True)
    logging.info('Average successfully calculated for del qty...')
    return average_del


def unrolled_OI_Calculator(expiry, month, year, flag):
    '''Function to compute unrolled OI sheet'''
    # unrolled OI calculator
    logging.info('Loading unrolled dumps master...')
    unrolled_today = pd.read_excel(master_dir + 'Unrolled_dumped.xlsx')
    unrolled_today = unrolled_today.set_index('SYMBOL')

    # average is appended in sorted manner
    stock_names = list(sorted(set(unrolled_today.index)))
    
    logging.info('Caclulate average of last three months for unrolled values...')
    unrolled_today['Date'] = unrolled_today.apply(lambda row: pd.to_datetime(row['Month']+' '+str(row['Year']) ), axis=1)
    date = pd.to_datetime(month+" "+year)
    
    average_unrolled = pd.DataFrame(average_calculator(expiry, 'Million_dollar', stock_names, unrolled_today, date),
                                    index=stock_names, columns=['Average'])
    average_unrolled.index.name = 'SYMBOL'
    if flag==0:
        return average_unrolled

    # read unrolled (million_dollar )from dumped data
    unrolled_OI_calc = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='Unrolled_rollovers')
    unrolled_OI_calc = unrolled_OI_calc.merge(average_unrolled, on='SYMBOL')


    # difference of today's million_unrolled and average of last 3 months
    unrolled_OI_calc['Difference'] = unrolled_OI_calc['Million_dollar'] - unrolled_OI_calc['Average']

    # Calculate percent for unrolled- avg/ avg
    unrolled_OI_calc['Percent'] = unrolled_OI_calc.apply(
        lambda row: round((row['Million_dollar'] - row['Average']) / row['Average'] * 100) if row['Average']!=0 else 0 , axis=1)

    # fetch del qty averages for last three months
    logging.info('Fetch del quantity average...')
    del_qty = del_qty_average()
    unrolled_OI_calc = unrolled_OI_calc.merge(del_qty, on='SYMBOL')
    unrolled_OI_calc['Ratio'] = unrolled_OI_calc['Million_dollar'] / unrolled_OI_calc['Del Avg']
    unrolled_OI_calc['Ratio'] = unrolled_OI_calc['Ratio'].apply(lambda x: round(x))

    # read master for fetching type of stock nifty, non-nifty and physical
    type_of_stock = pd.read_excel(master_dir + 'MasterData.xlsx')
    type_of_stock = type_of_stock[['SYMBOL','IsNifty','IsActiveFNO']]
    type_of_stock = type_of_stock[type_of_stock['IsActiveFNO']==True][['SYMBOL','IsNifty']]
    type_of_stock['TYPE'] = np.where(type_of_stock['IsNifty']==True, 'Nifty', 'Non Nifty')
    type_of_stock = type_of_stock[['SYMBOL','TYPE']]   
    
    
    unrolled_OI_calc = unrolled_OI_calc.merge(type_of_stock, on='SYMBOL')

    # pass False for top performing as sorting parameter (take only top stocks for unrolled)
    logging.info('Fetching top 7 stocks by percent...')
    unrolled_OI_top_stocks = sort_stocks(7, False, unrolled_OI_calc, 'Percent')

    # sort on ratio and million dollar values
    for key, values in unrolled_OI_top_stocks.items():
        unrolled_OI_top_stocks[key] = values.sort_values(ascending=False, by=['Ratio', 'Million_dollar']).head(4)
    logging.info('Returning top 4 stocks for unrolled OI')
    return unrolled_OI_top_stocks, average_unrolled


def increased_activity_calc(expiry, month, year):
    # increased Activity
    # read dumped value for SSF stucks for current expiry

    increased_activity = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='SSF_rollovers')
    increased_activity.set_index('SYMBOL', inplace=True)
    # read SSF stocks for T-1 expiry
    expiry_list = 'E-5 E-4 E-3 E-2 E-1 E'.split(' ')

    previous_day_expiry = pd.read_excel(master_dir + 'RollData.xlsx')
    logging.info('Fetching previous day expiry (T-1)...')
    previous_day_expiry = previous_day_expiry[
        (previous_day_expiry['Expiry'] == expiry_list[expiry_list.index(expiry) - 1]) & (
                previous_day_expiry['Month'] == month) &
        (previous_day_expiry['Year'] == int(year))][['SYMBOL', 'SSF_rollover']]
    previous_day_expiry.set_index('SYMBOL', inplace=True)

    # merge on symbol

    increased_activity = increased_activity.merge(previous_day_expiry, on='SYMBOL')
    increased_activity.columns = ['Expiry', 'BloomCode', 'Month', 'Year', 'SSF_rollover_T', 'SSF_rollover_T-1']

    # calculate difference Today - (T-1)

    increased_activity['Difference'] = increased_activity['SSF_rollover_T'] - increased_activity['SSF_rollover_T-1']
    increased_activity = increased_activity.sort_values(ascending=False, by='Difference').head(10)
    logging.info('Returning top 10 stocks showing incerased activity...')
    return increased_activity


def leading_lagging_sectors(expiry, month, year):
    '''Function to fetch leading and lagging sectors'''
    print month
    # read dumped indices
    dumped_indices = pd.read_excel(master_dir + 'Indices_dump.xlsx')
    dumped_indices = dumped_indices.groupby('Expiry').get_group(expiry)
    indices_names = list(sorted(set(dumped_indices['Indices'])))
    dumped_indices.set_index('Indices', inplace=True)
    dumped_indices.rename(index={'Technology':'tech','PSUBANKS':'PSU Banks'}, inplace=True)
    dic = {'Technology':'tech','PSUBANKS':'PSU Banks'}
    indices_names = [dic.get(n,n ) for n in indices_names]
    

    # Average of indicies for last 3 months
    logging.info('Calculating average for last three months for indicies...')
    dumped_indices['Date'] = dumped_indices.apply(lambda row: pd.to_datetime(row['Month']+' '+str(row['Year']) ), axis=1)
    date = pd.to_datetime(month+" "+year)
    
    avg_indices = pd.DataFrame(average_calculator(expiry, 'Roll', indices_names, dumped_indices, date),
                               index=indices_names, columns=['Average'])
    avg_indices.index.name = 'Indices'
    
    
    # consolidate data to already exisiting file
    from openpyxl import load_workbook
    book = load_workbook(output_dir+'Report.xlsx')
    
    writer = pd.ExcelWriter(output_dir+'Report.xlsx', engine='openpyxl') 
    writer.book = book
    
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    
    avg_indices.rename(index={'Technology':'tech','PSUBANKS':'PSU Banks'}, inplace=True)
    avg_indices.to_excel(writer, "Indices_avearge")
    writer.save()    
    
    #avg_indices.rename(index={'tech':'Technology','PSU Banks':'PSUBANKS'}, inplace=True)
   
    
    

    #avg_indices.to_excel(output_dir+'Average.xlsx',sheet_name = 'Indicies')

    indices_read = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='Indices_rollovers')
    indices_diff = indices_read.set_index('Indices')
    indices_diff = indices_diff.merge(avg_indices, on='Indices')
    indices_diff['Difference'] = indices_diff['Roll'] - indices_diff['Average']
    indices_diff['Difference'] = indices_diff.apply(lambda row: round(row['Difference'],2), axis=1)
    # Sectors to ignore
    
    exclude_list = ['Nifty', 'NiftyBank', 'Nifty Stocks', 'Non-Nifty Stocks', 'SSF', 'Overall', 'Real Estate',
                    'Diversified', 'chemicals', 'Media']
    logging.info('Excluding {0} sectors from the list...'.format(exclude_list))
    new_indices = indices_diff.drop(index=exclude_list)
    leading_sectors = []
    lagging_sectors = []
    # append leading and lagging sectors
    #topmost element

    if new_indices.sort_values('Difference', ascending = False)[:2].iloc[:-1,-1][0] == 0.0:
        logging.info('Dropping {0} sector since leading by {1}'.format(new_indices.sort_values('Difference', ascending = False)[:2].index[0],
                     new_indices.sort_values('Difference', ascending = False)[:2].iloc[:-1,-1][0]  ))
        
    
    else:
        leading_sectors.append(new_indices.sort_values('Difference', ascending = False)[:2].index[0])
    #2nd element in top stocks
    if new_indices.sort_values('Difference', ascending = False)[:2].iloc[-1:, -1][0] == 0.0:
        logging.info('Dropping {0} sector since leading by {1}'.format(new_indices.sort_values('Difference', ascending = False)[:2].index[1],
                     new_indices.sort_values('Difference', ascending = False)[:2].iloc[-1:,-1][0]  ))
    else:
        leading_sectors.append(new_indices.sort_values('Difference', ascending = False)[:2].index[1])
    #lowest stock
    if new_indices.sort_values('Difference')[:2].iloc[:-1,-1][0] == 0.0:
        logging.info('Dropping {0} sector since lagging by {1}'.format(new_indices.sort_values('Difference')[:2].index[0],
                     new_indices.sort_values('Difference')[:2].iloc[:-1,-1][0]  ))
    else:
        lagging_sectors.append(new_indices.sort_values('Difference')[:2].index[0])
    #2nd last stock
    if new_indices.sort_values('Difference')[:2].iloc[-1:, -1][0] == 0.0:
        logging.info('Dropping {0} sector since lagging by {1}'.format(new_indices.sort_values('Difference')[:2].index[1],
                     new_indices.sort_values('Difference')[:2].iloc[-1:,-1][0]  ))
    else:
        lagging_sectors.append(new_indices.sort_values('Difference')[:2].index[1])

    
    return leading_sectors, lagging_sectors, avg_indices


def print_final_report(expiry, avg_indices, increased_activity, leading_sectors, lagging_sectors,
                       grt_avg_top_stocks, grt_avg_low_stocks, unrolled_OI_top_stocks, file_name, month,year, average_unrolled):
    '''Function to print final report to excel'''


    # delete previous file 
    if os.path.exists(output_dir+expiry+month+year+"message.txt"):
        os.remove(output_dir+expiry+month+year+"message.txt")
    
    # write to text file 
    with open(output_dir+expiry+month+year+'message.txt','a') as myfile:
        
        
        

        myfile.write('----- Highlights at {0} -----\n'.format(expiry))
        
    
        if expiry == 'E':  
    
            
            for name in ['Nifty', 'NiftyBank', 'SSF', 'Nifty Stocks', 'Overall']:
                indices_read = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='Indices_rollovers')
                value = int(round(indices_read[indices_read['Indices'] == name]['Roll'].values[0]*100))
    
                avg = int(round(avg_indices.loc[name].values[0]*100))
                diff = value - avg
                if diff == 0:
                    myfile.write(' *{0} rolls ({1}%) are inline with the 3 month average of {2}%.\n\n '.format(
                            name, value , avg ))
                    
                
                elif diff > 0:
                    myfile.write(' *{0} rolls ({1}%) were {2} percentage points higher than the 3 month average of {3}%. \n\n'.format(
                        name, value , diff , avg ))
                    
                   
                
                    
                else:
                    myfile.write(' *{0} rolls ({1}%) were {2} percentage points lower than the 3 month average of {3}%. \n\n'.format(
                        name, value , diff * -1, avg * 1))
                   


                   
    
            if expiry in ['E-5', 'E-4', 'E-3']:
                pass
            else:
                myfile.write(' *Increased roll activity was seen in {0} \n\n'.format(
                    ', '.join(str(x) for x in increased_activity['BloomCode'].values)))      
              
    
            if len(leading_sectors) == 0 and len(lagging_sectors) == 0:
                pass
            elif len(leading_sectors) == 1 and len(lagging_sectors) == 0:
                myfile.write('*Rolls in {0} lead their 3 month average\n\n '.format(leading_sectors.pop()))
               
               
            elif len(leading_sectors) == 0 and len(lagging_sectors) == 1:
                myfile.write('*Rolls in {0} lagged their 3 month average\n\n '.format(lagging_sectors.pop()))
               
               
            elif len(leading_sectors) == 2 and len(lagging_sectors) == 1:
                myfile.write('*Rolls in {0} and {1} lead their 3 month average, while those in {2} lagged their 3 month average \n\n'.format(leading_sectors.pop(),
                                                                                                                                                                leading_sectors.pop(),
                                                                                                                                                                lagging_sectors.pop()))
               
               
            elif len(leading_sectors) == 1 and len(lagging_sectors) == 2:
                myfile.write('*Rolls in {0} lead their 3 month average, while those in {1} and {2} lagged their 3 month average\n\n '.format(leading_sectors.pop(),
                                                                                                                                                                 lagging_sectors.pop(),
                                                                                                                                                                 lagging_sectors.pop()))
                
               
            elif len(leading_sectors) == 2 and len(lagging_sectors) == 2:
                myfile.write('*Rolls in {0} and {1} lead their 3 month average, while those in {2} and {3} lagged their 3 month average \n\n'.format(leading_sectors.pop(),
                                                                                                                                                                         leading_sectors.pop(),
                                                                                                                                                                         lagging_sectors.pop(),
                                                                                                                                                                         lagging_sectors.pop()))
               
                   
    
            myfile.write(' *Stocks with rollovers higher than their 3 month average: ')    
            
    
            i = 1
    
            data = OrderedDict(reversed(list(grt_avg_top_stocks.items())))
            while True:
                for key, values in data.iteritems():
                    myfile.write('\n {0} \n'.format(key))
                    try:
                        if values=='None':
                            myfile.write("None")
                    except:                        
                        stocks = []
                        for name in values['BloomCode'].values:
                            stocks.append("{0}({1}% Vs {2}%)".format(str(name),
                                                             int(round(
                                                                 values.set_index('BloomCode').loc[name][['SSF_rollover']][
                                                                     0] * 100)),int(round(
                                                                 values.set_index('BloomCode').loc[name][['Average']][
                                                                     0] * 100)) ))
                        myfile.write(', '.join(stocks))
                    
                i += 1
                if i > 2:
                    break
                data = OrderedDict(reversed(list(grt_avg_low_stocks.items())))
              
                myfile.write('\n\n *Stocks with rollovers lower than their 3 month average: ')
               
           
            #myfile.write("\n\n\n This is an auto generated email, please do not reply to it !")
            myfile.close()
            #pd.DataFrame(print_data).to_excel(output_dir + file_name+ 'Final.xlsx', sheet_name='Report', header=None, index=None)
          
            #Email_func.expiry_email(month,year,expiry)
    
        else:   
            
    
            for name in ['Nifty', 'NiftyBank', 'SSF', 'Nifty Stocks', 'Overall']:
                indices_read = pd.read_excel(output_dir + 'Report.xlsx', sheet_name='Indices_rollovers')
                value = int(round(indices_read[indices_read['Indices'] == name]['Roll'].values[0]*100))
    
                avg = int(round(avg_indices.loc[name].values[0]*100))
                diff = value - avg
                if diff == 0:
                    myfile.write(' *{0} rolls ({1}%) are inline with the 3 month average of {2}%. \n\n'.format(
                        name, value , avg ))                       
                
                
                elif diff > 0:
                    myfile.write(' *{0} rolls ({1}%) are {2} percentage points higher than the 3 month average of {3}%. \n\n'.format(
                        name, value , diff , avg ))
                    
                                       
                else:
                    myfile.write(' *{0} rolls ({1}%) are {2} percentage points lower than the 3 month average of {3}%.\n\n '.format(
                        name, value , diff * -1, avg ))
                   
                       
            if expiry in ['E-5', 'E-4', 'E-3']:
                pass
            else:
                myfile.write(' *Increased roll activity was seen in {0}\n\n '.format(
                    ', '.join(str(x) for x in increased_activity['BloomCode'].values)))
               
                
    
            # leading and lagging print stmt
            if len(leading_sectors) == 0 and len(lagging_sectors) == 0:
                pass
            elif len(leading_sectors) == 1 and len(lagging_sectors) == 0:
                myfile.write('*Rolls in {0} are leading their 3 month average\n\n '.format(leading_sectors.pop()))
               
               
            elif len(leading_sectors) == 0 and len(lagging_sectors) == 1:
                myfile.write('*Rolls in {0} are lagging their 3 month average \n\n'.format(lagging_sectors.pop()))
               
                
            elif len(leading_sectors) == 2 and len(lagging_sectors) == 1:
                myfile.write('*Rolls in {0} and {1} are leading their 3 month average, while those in {2} are lagging their 3 month average\n\n '.format(leading_sectors.pop(),
                                                                                                                                                                            leading_sectors.pop(),
                                                                                                                                                                            lagging_sectors.pop()))
              
               
            elif len(leading_sectors) == 1 and len(lagging_sectors) == 2:
                myfile.write('*Rolls in {0} are leading their 3 month average, while those in {1} and {2} are lagging their 3 month average\n\n '.format(leading_sectors.pop(),
                                                                                                                                                                             lagging_sectors.pop(),
                                                                                                                                                                             lagging_sectors.pop()))
              
               
            elif len(leading_sectors) == 2 and len(lagging_sectors) == 2:
                myfile.write('*Rolls in {0} and {1} are leading their 3 month average, while those in {2} and {3} are lagging their 3 month average\n\n '.format(leading_sectors.pop(),
                                                                                                                                                                                     leading_sectors.pop(),
                                                                                                                                                                                     lagging_sectors.pop(),
                                                                                                                                                                                     lagging_sectors.pop()))
                
                   
            myfile.write(' *Stocks with rollovers higher than their 3 month average: ')          
            i = 1   
            data = OrderedDict(reversed(list(grt_avg_top_stocks.items())))
            while True:
                for key, values in data.iteritems():
                    myfile.write('\n {0}\n '.format(key))
                    try:
                        if values=='None':
                            myfile.write("None")
                    except:                   
                        stocks = []
                        for name in values['BloomCode'].values:
                            stocks.append("{0}({1}% Vs {2}%)".format(str(name),
                                                             int(round(
                                                                 values.set_index('BloomCode').loc[name][['SSF_rollover']][
                                                                     0] * 100)),int(round(
                                                                 values.set_index('BloomCode').loc[name][['Average']][
                                                                     0] * 100)) ))
                        myfile.write(', '.join(stocks))
                   
                i += 1
                if i > 2:
                    break
                data = OrderedDict(reversed(list(grt_avg_low_stocks.items())))
            
                myfile.write('\n\n *Stocks with rollovers lower than their 3 month average: ')
               
    
         
            myfile.write('\n\n *SSF with high unrolled OI compared to 3 mn avg: (OI in million dollars and Ratio of OI to 3mn avg of expiry day delivery )')
           
          
    
            for key, values in OrderedDict(reversed(list(unrolled_OI_top_stocks.items()))).iteritems():
                myfile.write('\n {0}\n '.format(key))
               
                stocks = []
                for name in values['BloomCode']:
                    stocks.append(
                        '{0}({1}/{2}x)'.format(name, values.set_index('BloomCode').loc[name]['Million_dollar'],
                                               values.set_index('BloomCode').loc[name]['Ratio']))
                myfile.write(', '.join(stocks))
               
            logging.info('Final reprot successfully generated for {0} '.format(file_name))
            
            #book.save(output_dir + file_name + 'Final.xlsx')
            #myfile.write("\n\n\n This is an auto generated email, please do not reply to it !")
            myfile.close()
            #pd.DataFrame(print_data).to_excel(output_dir + file_name+ 'Final.xlsx', sheet_name='Report', header=None, index=None)
          
            #Email_func.expiry_email(month,year,expiry)
            #os.remove(email_dir+"message.txt")            
            
        #pd.DataFrame(print_data).to_excel(output_dir + file_name+'Final.xlsx', sheet_name='Report', header=None, index=None)
             
        shutil.copyfile(output_dir+expiry+month+year+"message.txt", email_dir+'message.txt')
        shutil.copyfile(output_dir+expiry+month+year+"message.txt", market_snapshot+'message.txt')
        # ssf rollovers 
        df = pd.read_excel(output_dir+'Report.xlsx', sheet_name='SSF_rollovers')
        #df['SSF_rollover'] = df['SSF_rollover']*100
        df1 = pd.read_excel(output_dir+'Report.xlsx', sheet_name='SSF_avearge')
        #df1['Average'] = df1['Average']*100
        # merge ssf rollovers and average 
        df = df.merge(df1, on='SYMBOL', how = 'left')
        
        # indices rollovers
        df2 = pd.read_excel(output_dir+'Report.xlsx', sheet_name='Indices_rollovers')
        #df2['Roll'] = df2['Roll']*100
        df3 = pd.read_excel(output_dir+'Report.xlsx', sheet_name='Indices_avearge')
        #df3['Average'] = df3['Average']*100
        # merge 
        df2 = df2.merge(df3, on='Indices', how = 'left')        
        # unrolled rollovers
        df4 = pd.read_excel(output_dir+'Report.xlsx', sheet_name='Unrolled_rollovers')
        
        # write dataframe to excel
        writer = pd.ExcelWriter(output_dir+'Report.xlsx', engine='openpyxl')        
        df.to_excel(writer, sheet_name='SSF_rollovers', index=False)
        df2.to_excel(writer, sheet_name='Indices_rollovers', index=False)
        df4.to_excel(writer, sheet_name='Unrolled_rollovers', index=False)
        writer.save()
        writer.close()
        
        
        ####reorder code####
        symbol_order = pd.read_csv(master_dir+'symbol_order.csv', header=None).fillna(0)
        sectors = pd.read_excel(master_dir+'MasterData.xlsx')[['SYMBOL','Sector']]
        sectors.rename(columns={'SYMBOL':'Stock'}, inplace=True)

        # ssf rollovers
        df1 = pd.read_excel(output_dir + 'Report.xlsx', sheet_name="SSF_rollovers")
        expiry = df1['Expiry'][0]
        df1.rename(columns={'SYMBOL':'Stock','BloomCode':'Berg','SSF_rollover':expiry, 'Average':'Avg'}, inplace=True)
        df1.drop(columns=['Expiry','Month','Year'], inplace=True)
        ssf_symbols = symbol_order[symbol_order[0]=='ssf']
        ssf_symbols.rename(columns={1:'Stock'}, inplace=True)
        ssf_result = ssf_symbols.merge(df1, on='Stock', how='left', sort=False).drop(columns=0)        
        ssf_result = ssf_result.merge(sectors, on='Stock', how='left')
        ssf_result = ssf_result[['Sector','Stock','Berg',expiry,'Avg']]
        
        
        
        
        # indices rollovers
        df1 = pd.read_excel(output_dir+'Report.xlsx', sheet_name="Indices_rollovers")
        expiry = df1['Expiry'][0]
        df1.rename(columns={'Indices':'  ','Roll':expiry, 'Average':'Avg'}, inplace=True)
        df1.drop(columns=['Expiry','Month','Year'], inplace=True)
        
        sector_order = symbol_order[symbol_order[0]=='sectors']
        indices_result = sector_order.rename(columns={1:'  '}).merge(df1 ,on='  ', 
                                                                     how ='left', sort=False).drop(columns=[0]).fillna(' ').replace(0,'')
        
        # unrolled OI ordering 
        df1 = pd.read_excel(output_dir+'Report.xlsx', sheet_name="Unrolled_rollovers")
        average_unrolled.rename(columns={'Average':'Avg'}, inplace=True)
        df1 = df1.merge(average_unrolled, on='SYMBOL', how='left', sort=False)
        expiry = df1['Expiry'][0]
        df1.rename(columns={'SYMBOL':'Stock','BloomCode':'Berg','Million_dollar':'Unrolled OI $mn'}, inplace=True)
        df1.drop(columns=['Expiry','Month','Year'], inplace=True)
        unrolled_symbols = symbol_order[symbol_order[0]=='del']
        unrolled_symbols.rename(columns={1:'Stock'}, inplace=True)
        unrolled_result = unrolled_symbols.merge(df1, on='Stock', how='left', sort=False).drop(columns=0)        
        unrolled_result = unrolled_result.merge(sectors, on='Stock', how='left')
        unrolled_result = unrolled_result[['Sector','Stock','Berg','Unrolled OI $mn','Avg']]
        
        
        
        
        writer = pd.ExcelWriter(output_dir+'Report.xlsx', )
        ssf_result.to_excel(writer, startcol=0,startrow=0, index=False, sheet_name=expiry)
        indices_result.to_excel(writer, startcol=6, startrow=0, index=False, sheet_name=expiry)
        unrolled_result.to_excel(writer, startcol=10, startrow=0, index=False, sheet_name=expiry)
        writer.save()
        writer.close()       
        
        
        ##############    
        
        
        
        shutil.copyfile(output_dir+'Report.xlsx', email_dir+'Expiry_roll_data.xlsx')
        os.rename(output_dir+'Report.xlsx', output_dir+'Report_{}.xlsx'.format(datetime.datetime.today().date()))
        os.remove(download_dir+'fo'+file_name+'bhav.csv')        
       



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       **to_excel_kwargs):
 
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(master_dir+filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(master_dir+filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except Exception as e:
        # file does not exist yet, we will create it
        print e

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()



# Remove duplicate rows from master data
def drop_duplicate_rows(filename):
    '''Function to drop duplicate rows from master dumps'''
    logging.info('Removing duplicates from master file {0}...'.format(filename))
    
    df = pd.read_excel(master_dir+filename, sheet_name= 'Sheet1')
    df.drop_duplicates(keep='first', inplace=True)
    df.to_excel(master_dir+filename, index = False)
